# Copyright (c) Cosmo Tech corporation.
# Licensed under the MIT license.
import csv
import io
import json
import os
import tempfile
from typing import Union

from azure.digitaltwins.core import DigitalTwinsClient
from azure.identity import DefaultAzureCredential

import cosmotech_api
from cosmotech_api.api.dataset_api import DatasetApi
from cosmotech_api.api.scenario_api import ScenarioApi
from cosmotech_api.api.workspace_api import WorkspaceApi

from openpyxl import load_workbook


class ScenarioDownloader:

    def __init__(self, workspace_id: str, organization_id: str):
        self.credentials = DefaultAzureCredential()
        scope = os.environ.get("COSMOTECH_API_SCOPE")
        token = self.credentials.get_token(scope)

        self.configuration = cosmotech_api.Configuration(
            host=os.environ.get("COSMOTECH_API_HOST"),
            discard_unknown_keys=True,
            access_token=token.token
        )

        self.workspace_id = workspace_id
        self.organization_id = organization_id

    def get_scenario_data(self, scenario_id: str):
            with cosmotech_api.ApiClient(self.configuration) as api_client:
                api_instance = ScenarioApi(api_client)
                scenario_data = api_instance.find_scenario_by_id(organization_id=self.organization_id,
                                                                 workspace_id=self.workspace_id,
                                                                 scenario_id=scenario_id)
            return scenario_data

    def download_dataset(self, dataset_id: str) -> (str, str, Union[str, None]):
        with cosmotech_api.ApiClient(self.configuration) as api_client:
            api_instance = DatasetApi(api_client)

            dataset = api_instance.find_dataset_by_id(organization_id=self.organization_id,
                                                      dataset_id=dataset_id)

            is_adt = 'AZURE_DIGITAL_TWINS_URL' in dataset['connector']['parameters_values']

            if is_adt:
                return {"type": 'adt',
                        "content": self._download_adt_content(adt_adress=dataset['connector']['parameters_values'][
                            'AZURE_DIGITAL_TWINS_URL']), "name": dataset['name']}
            else:
                _file_name = dataset['connector']['parameters_values']['AZURE_STORAGE_CONTAINER_BLOB_PREFIX'].replace(
                    '%WORKSPACE_FILE%/', '')
                return {"type": _file_name.split('.')[-1], "content": self._download_file(_file_name),
                        "name": dataset['name']}

    def _download_file(self, file_name: str):
        tmp_dataset_dir = tempfile.mkdtemp()
        target_file = os.path.join(tmp_dataset_dir, file_name.split('/')[-1])
        with cosmotech_api.ApiClient(self.configuration) as api_client:
            api_ws = WorkspaceApi(api_client)

            dl_file = api_ws.download_workspace_file(organization_id=self.organization_id,
                                                     workspace_id=self.workspace_id,
                                                     file_name=file_name)

            with open(target_file, "wb") as tmp_file:
                tmp_file.write(dl_file.read())
            content = dict()
            if ".xls" in file_name:
                wb = load_workbook(target_file, data_only=True)
                for sheet_name in wb.sheetnames:
                    sheet = wb[sheet_name]
                    content[sheet_name] = list()
                    headers = next(sheet.iter_rows(max_row=1, values_only=True))

                    def item(_row: tuple) -> dict:
                        return {k: v for k, v in zip(headers, _row)}

                    for r in sheet.iter_rows(min_row=2, values_only=True):
                        row = item(r)
                        new_row = dict()
                        for key, value in row.items():
                            try:
                                converted_value = json.load(io.StringIO(value))
                            except (json.decoder.JSONDecodeError, TypeError):
                                converted_value = value
                            if converted_value is not None:
                                new_row[key] = converted_value
                        if new_row:
                            content[sheet_name].append(new_row)
            elif ".csv" in file_name:
                with open(target_file, "r") as file:
                    # Read every file in the input folder
                    current_filename = os.path.basename(target_file)[:-len(".csv")]
                    content[current_filename] = list()
                    for row in csv.DictReader(file):
                        new_row = dict()
                        for key, value in row.items():
                            try:
                                # Try to convert any json row to dict object
                                converted_value = json.load(io.StringIO(value))
                            except json.decoder.JSONDecodeError:
                                converted_value = value
                            if converted_value == '':
                                converted_value = None
                            if converted_value is not None:
                                new_row[key] = converted_value
                        content[current_filename].append(new_row)
            elif ".json" in file_name:
                with open(target_file, "r") as _file:
                    content = json.load(_file)
            else:
                with open(target_file, "r") as _file:
                    content = "\n".join(line for line in _file)
        return content

    def _download_adt_content(self, adt_adress: str) -> dict:
        client = DigitalTwinsClient(adt_adress, self.credentials)
        query_expression = 'SELECT * FROM digitaltwins'
        query_result = client.query_twins(query_expression)
        json_content = dict()
        for twin in query_result:
            entity_type = twin.get('$metadata').get('$model').split(':')[-1].split(';')[0]
            t_content = {k: v for k, v in twin.items()}
            t_content['id'] = t_content['$dtId']
            for k in twin.keys():
                if k[0] == '$':
                    del t_content[k]
            json_content.setdefault(entity_type, [])
            json_content[entity_type].append(t_content)

        relations_query = 'SELECT * FROM relationships'
        query_result = client.query_twins(relations_query)
        for relation in query_result:
            tr = {
                "$relationshipId": "id",
                "$sourceId": "source",
                "$targetId": "target"
            }
            r_content = {k: v for k, v in relation.items()}
            for k, v in tr.items():
                r_content[v] = r_content[k]
            for k in relation.keys():
                if k[0] == '$':
                    del r_content[k]
            json_content.setdefault(relation['$relationshipName'], [])
            json_content[relation['$relationshipName']].append(r_content)

        return json_content

    def get_all_parameters(self, scenario_id) -> dict:
        scenario_data = self.get_scenario_data(scenario_id=scenario_id)
        content = dict()
        for parameter in scenario_data['parameters_values']:
            content[parameter['parameter_id']] = parameter['value']
        return content

    def get_all_datasets(self, scenario_id: str) -> dict:
        scenario_data = self.get_scenario_data(scenario_id=scenario_id)

        datasets = scenario_data['dataset_list']

        content = dict()
        for dataset_id in datasets:
            content[dataset_id] = self.download_dataset(dataset_id)

        for parameter in scenario_data['parameters_values']:
            if parameter['var_type'] == '%DATASETID%':
                dataset_id = parameter['value']
                content[dataset_id] = self.download_dataset(dataset_id)

        return content